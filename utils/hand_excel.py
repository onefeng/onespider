# !/usr/bin/python
# -*- coding: UTF-8 -*-
import openpyxl

class HandExcel(object):

    @staticmethod
    def insert_one(filepath, sheetname, data):
        """
        追加文件
        :param filepath:
        :param sheetname:
        :param datas:
        :return:
        """
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheetname]
        # 待填充数据
        ws.append(data)
        savename = filepath
        wb.save(savename)

    @staticmethod
    def insert(filepath, sheetname,datas):
        """
        追加文件
        :param filepath:
        :param sheetname:
        :param datas:
        :return:
        """
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheetname]
        # 待填充数据
        for data in datas:
            ws.append(data)
        savename = filepath
        wb.save(savename)
    def test_read_excel(self, tablename, sheetname,star_row=2):
        """
        读取文件
        :param tablename:
        :param sheetname:
        :param star_row: 因为第一行是标题 默认从第2行读
        :return:
        """
        workbook = openpyxl.load_workbook(tablename)
        worksheet = workbook[sheetname]
        row = worksheet.max_row
        col = worksheet.max_column
        value_list=[]
        for i in range(star_row, row+1):
            value_row=[]
            for j in range(1,col+1):
                unit = worksheet.cell(i, j).value
                value_row.append(unit)
            value_list.append(value_row)
        return value_list

    @staticmethod
    def read_excel(tablename, sheetname,star_row=2):
        """
        读取文件
        :param tablename:
        :param sheetname:
        :param star_row: 因为第一行是标题 默认从第2行读
        :return:
        """
        workbook = openpyxl.load_workbook(tablename)
        worksheet = workbook[sheetname]
        row = worksheet.max_row
        col = worksheet.max_column
        value_list=[]
        for i in range(star_row, row+1):
            value_row=[]
            for j in range(1,col+1):
                unit = worksheet.cell(i, j).value
                value_row.append(unit)
            value_list.append(value_row)
        return value_list
    #写入文件
    @staticmethod
    def write_excel(name, sheet_name, items):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        for i in range(0,len(items)):
            for j in range(0,len(items[i])):
                worksheet.cell(i+1, j+1, items[i][j])
        workbook.save(name)

    #修改已有文件
    def revise_excel(self,tablename,sheetname,items):
        workbook = openpyxl.load_workbook(tablename)
        worksheet = workbook[sheetname]
        #worksheet.insert_cols(9)
        for i in range(0, len(items)):
            for j in range(0, len(items[i])):
                worksheet.cell(i + 1, j + 1, items[i][j])
        workbook.save(tablename)

if __name__ == '__main__':
    pass
    datas=HandExcel.read_excel('origin.xlsx','2020q1')
    for data in datas:
        print(len(str(data[2])))
        print(len(str(data[2]))==4)
